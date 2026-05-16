import openpyxl, json, sys
from pathlib import Path

wb = openpyxl.load_workbook(r'C:\Users\alrev\WebDev\tenni\Fantasy Tennis 2025.xlsx', data_only=True)
out = Path(r'C:\Users\alrev\WebDev\tenni\data\sheet_dump')
out.mkdir(parents=True, exist_ok=True)

for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() != '' for c in row):
            rows.append([('' if c is None else str(c)) for c in row])
    safe = name.replace('/','_').replace(',','_').replace(' ','_')
    (out / f'{safe}.json').write_text(json.dumps(rows, ensure_ascii=False, indent=1))
print('done', len(wb.sheetnames))
